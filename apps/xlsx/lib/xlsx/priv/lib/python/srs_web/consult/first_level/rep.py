#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
#from importlib import reload
#reload(sys)
#from patch import patch_worksheet
import time
import datetime
import json
# import pytz
# from bson import ObjectId
# from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl import load_workbook
# from bson.codec_options import CodecOptions
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Variables
#-------------------------------------------------------------------------------
letters = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']
cluesNames = []

# -- Estilos

def subtitleStyle(cell):
    font = Font(name='Arial', bold=False, color='000000', size=24)

    cell.font = font

    return cell

def diagnosisStyle(cell):
    font = Font(name='Arial', bold=False, color='000000', size=30)

    cell.font = font

    return cell

def diagnosisTitleStyle(cell):
    font = Font(name='Arial', bold=False, color='000000', size=22)

    cell.font = font

    return cell

# Funciones Reporte
#-------------------------------------------------------------------------------

def generateReport (consults,data):

    wb = load_workbook(data.get('pathTemplate'))



    for i in range(0, 3):

        wb.active = i
        ws = wb.active

        imgLogo = Image(data.get('logo'))
        ws.add_image(imgLogo, 'D4')

        firstRow = 13
        secondRow = 15
        # if get(data,['params','clueName']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('D'))).value = get(data,['params','clueName'])
        #
        # if get(data,['params','startDate','day']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('M'))).value = get(data,['params','startDate','day'])
        #
        # if get(data,['params','startDate','month']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('O'))).value = get(data,['params','startDate','month'])
        #
        # if get(data,['params','startDate','year']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('S'))).value = get(data,['params','startDate','year'])
        #
        # if get(data,['params','endDate','day']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('U'))).value = get(data,['params','endDate','day'])
        #
        # if get(data,['params','endDate','month']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('W'))).value = get(data,['params','endDate','month'])
        #
        # if get(data,['params','endDate','year']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('AA'))).value = get(data,['params','endDate','year'])
        #
        # if get(data,['params','clue']):
        #     subtitleStyle(ws.cell(row = firstRow, column = letters.index('AD'))).value = get(data,['params','clue'])
        #
        # if get(data,['params','location']):
        #     subtitleStyle(ws.cell(row = secondRow, column = letters.index('D'))).value = get(data,['params','location'])
        #
        # if get(data,['params','municipality']):
        #     subtitleStyle(ws.cell(row = secondRow, column = letters.index('J'))).value = get(data,['params','municipality'])
        #
        # if get(data,['params','jurisdiction']):
        #     subtitleStyle(ws.cell(row = secondRow, column = letters.index('T'))).value = get(data,['params','jurisdiction'])
        #
        # if get(data,['params','state']):
        #     subtitleStyle(ws.cell(row = secondRow, column = letters.index('AC'))).value = get(data,['params','state'])
        #
        # if get(data,['params','institution_name']):
        #     subtitleStyle(ws.cell(row = 17, column = letters.index('J'))).value = get(data,['params','institution_name'])

    wb.active = 0
    ws = wb.active

    group1 = consults.get('group1')
    group2 = consults.get('group2')
    startRow = 22
    startColumn = 6
    startRow2 = 102

    for diagnostic in group1:
        if startRow == 51 :
            startRow = 57
        if startRow == 95 :
            wb.active = 1
            ws = wb.active
            startRow = 22

        diagnosisTitleStyle(ws.cell(row = startRow, column = letters.index('D'))).value = get(diagnostic,['name'])

        groupAges = get(diagnostic,['groupAges'])

        for group in groupAges:
            if(startColumn == 30):
                startColumn = 6

            diagnosisStyle(ws.cell(row = startRow, column = startColumn)).value = get(group,['mens'])
            startColumn += 1

            diagnosisStyle(ws.cell(row = startRow, column = startColumn)).value = get(group,['womens'])
            startColumn += 1

        startRow += 1


    for diagnostic2 in group2:
        if startRow == 51 :
            startRow = 57

        if startRow == 90 :
            wb.active = 2
            ws = wb.active
            startRow = 22

        diagnosisTitleStyle(ws.cell(row = startRow, column = letters.index('D'))).value = get(diagnostic2,['name'])

        groupAges2 = get(diagnostic2,['groupAges'])

        for item2 in groupAges2:
            if(startColumn == 30):
                startColumn = 6

            diagnosisStyle(ws.cell(row = startRow, column = startColumn)).value = get(item2,['mens'])
            startColumn += 1

            diagnosisStyle(ws.cell(row = startRow, column = startColumn)).value = get(item2,['womens'])
            startColumn += 1

        startRow += 1

    wb.active = 0
    #wb.save(data.get('path'))
    wb.save("/home/lety/file123.xlsx")
    print ('success')


# Utilidades
#-------------------------------------------------------------------------------
def get(obj,list):
    for item in list:
        if obj.get(item) :
            obj = obj.get(item)
        else :
            return ""
    return obj;


# Iniciar reporte
#-------------------------------------------------------------------------------

def initrep (str_json):
    Json = json.loads(str_json)
    data = Json.get("data")
    consults = Json.get("consults")

    generateReport(consults,data)
